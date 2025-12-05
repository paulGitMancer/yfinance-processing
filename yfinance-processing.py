# -*- coding: utf-8 -*-
"""
Created on Tue Nov 25 17:58:18 2025

@author: paulGitMancer
"""

#----------------------------------------------------------------------------#
#-------------------------- TOP 200 STOCK ANALYSIS --------------------------#
#------------------------------ paulGitMancer -------------------------------#
#----------------------------------------------------------------------------#

#----------------------------------------------------------------------------#
#--------------------------------- README: ----------------------------------#
#----- This Python script performs an automated analysis of 200 major -------#
#----- stocks. The script reads a list of ticker symbols from an Excel ------#
#----- file designated below (line 38) and uses yfinance to retrieve the ----#
#----- current stock data. The script will also perform a few numerical -----#
#----- calculations. The data is then written to a new workbook and some ----#
#----- minor formatting is performed. The final output workbook is saved ----#
#-------------- to the file path designated below (line 166). ---------------#
#----------------------------------------------------------------------------#

#-----------------------------------------#
#------------ IMPORT LIBRARIES -----------#
#-----------------------------------------#

import yfinance as yf                                                           
import openpyxl                                                                 
from openpyxl.styles import Font
from openpyxl.styles import Alignment          

#------------------------------------------#
#----------- LOAD TICKERS FILE ------------#
#------------------------------------------#

wb_input = openpyxl.load_workbook \
    ('C://Tickers_File.xlsx')                                                   # Opens the workbook containing top 200 tickers.
sheet_input = wb_input.active                                                   # Selects the active sheet from the workbook.

#------------------------------------------#
#--------- CREATE OUTPUT WORKBOOK ---------#
#------------------------------------------#

wb_output = openpyxl.Workbook()                                                 # Creates a new workbook.
sheet_output = wb_output.active                                                 # Selects the active sheet from the new workbook.
sheet_output.title = "Stock Analysis"                                           # Renames the sheet.

#------------------------------------------#
#------------- CREATE HEADERS -------------#
#------------------------------------------#

headers = [                                                                     # Identifies column header names using a list.
    "Ticker", "Company Name", "Sector", "Current Price", \
    "Market Cap", "52 Week Low", "52 Week High", "% Above Low", \
    "% Below High", "Market Cap Category"]
    
for col_number, header in enumerate(headers,start=1):                           # Loop using enumerate to pair column numbers with 'headers' list. Starts at column one.
    cell_obj = sheet_output.cell(row=1,column=col_number)                       # Identifying output cell using 'col_number' variable.
    cell_obj.value = header                                                     # Writes header names into cells.
    cell_obj.font = Font(bold=True, size=14)                                    # Sets header font to bold and increases font size to 14.
        
#------------------------------------------#
#--------- TICKER DATA PROCESSING ---------#
#------------------------------------------#

output_row = 2                                                                  # Identifies row to begin writing data.
                                                                  
for input_row in range(1,sheet_input.max_row+1):                                # Loops through all rows in the input file,
    ticker = sheet_input.cell(row=input_row,column=1).value                     # reads the ticker reference in column A,
    stock_obj = yf.Ticker(ticker)                                               # and creates a yfinance ticker for that reference.
       
    try:                                                                        # Without 'try' and 'except', HTTP errors often crash the script.
        stock_info = stock_obj.info                                             # I had to do a bit of research to figure this one out.
    except Exception:                                                           # Creates an exception.
        print("Error processing data for:", ticker)                             # If HTTP error occurs, identifies which ticker timed out.
        continue                                                                # Script continues despite HTTP errors.

#------------------------------------------#
#---------- STORE YFINANCE DATA -----------#
#------------------------------------------#      
                                                                                # Using 'get' to store data. Without 'get', KeyError occurs.
    company_name = stock_info.get("shortName", "N/A")                           # Using 'company_name' variable to store company short name.
    company_sector = stock_info.get("sector", "N/A")                            # Using 'company_sector' variable to store company sector.
    current_price = stock_info.get("currentPrice", 0)                           # Using 'current_price' variable to store current price.
    market_cap = stock_info.get("marketCap", 0)                                 # Using 'market_cap' variable to store market cap.
    fifty_two_week_low = stock_info.get("fiftyTwoWeekLow", 0)                   # Using 'fifty_two_week_low' variable to store 52-week low.
    fifty_two_week_high = stock_info.get("fiftyTwoWeekHigh", 0)                 # Using 'fifty_two_week_high' variable to store 52-week high.    

#------------------------------------------#
#----- NUMERICAL ANALYSIS CALCULATIONS ----#
#------------------------------------------#
    
    if fifty_two_week_low != 0:                                                 # Without the 'if' and 'else' statements, ZeroDivisionError occurs.
        percent_above_low = (                                                   # Numerical calculation to determine how far above the 52-week
            current_price - fifty_two_week_low) / fifty_two_week_low            # low the current price is.
    else:
        percent_above_low = 0
        
    if fifty_two_week_high != 0:                                                # Without the 'if' and 'else' statements, ZeroDivisionError occurs.
        percent_below_high = (                                                  # Numerical calculation to determine how far below the 52-week
            fifty_two_week_high - current_price) / fifty_two_week_high          # high the current price is.
    else:
        percent_below_high = 0
        
    if market_cap >=100_000_000_000:                                            # Simple 'if' statement to categorize companies by market-cap size.      
        mcap_category = "Large Market Cap"                                      # I couldn't figure out how to calculate one-year return without
    elif market_cap >=50_000_000_000:                                           # historical data, so I opted for this instead. 
        mcap_category = "Mid Market Cap"                                         
    else:
        mcap_category = "Small Market Cap"

#------------------------------------------#
#----- WRITE VALUES INTO NEW WORKBOOK -----#
#------------------------------------------#

    output_values = [                                                           # 'output_values' list references variables from yfinance section above.
        ticker, company_name, company_sector, current_price, market_cap, \
        fifty_two_week_low, fifty_two_week_high, percent_above_low, \
        percent_below_high, mcap_category]
        
    for col_number, value in enumerate(output_values, start=1):                 # Loop using enumerate to pair column numbers with 'output_values' list.
        sheet_output.cell(row=output_row, column=col_number).value = value      # Writes values into cells.
        
    output_row += 1                                                             # Moves to the next 'output_row' after each row of data is written.

#------------------------------------------#
#------------- EXCEL FORMATTING -----------#
#------------------------------------------#

col_currentprice = 4                                                            # Using 'col_currentprice' variable to store column number.
col_mcap = 5                                                                    # Using 'col_mcap' variable to store column number.
col_52_low = 6                                                                  # Using 'col_52_low' variable to store column number.
col_52_high = 7                                                                 # Using 'col_52_high' variable to store column number.
col_pal = 8                                                                     # Using 'col_pal' variable to store column number.
col_pbh = 9                                                                     # Using 'col_pbh' variable to store column number.

for row_number in range(2, sheet_output.max_row+1):                             # Loops through each row in the output file (skips header row).
    sheet_output.cell(      
        row=row_number, column=col_currentprice).number_format = '$#,##0.00'    # Changes current price column format to currency.
    sheet_output.cell(
        row=row_number, column=col_52_low).number_format = '$#,##0.00'          # Changes 52-week low column format to currency.
    sheet_output.cell(
        row=row_number, column=col_52_high).number_format = '$#,##0.00'         # Changes 52-week high column format to currency.

for row_number in range(2, sheet_output.max_row+1):                             # Loops through each row in the output file (skips header row).
    sheet_output.cell(
        row=row_number, column=col_mcap).number_format = '$#,##0'               # Changes market cap column format to currency instead of scientific notation.

for row_number in range(2, sheet_output.max_row+1):                             # Loops through each row in the output file (skips header row).
    sheet_output.cell(
        row=row_number, column=col_pal).number_format = '0.00%'                 # Changes percent above low column format to percentage.
    sheet_output.cell(
        row=row_number, column=col_pbh).number_format = '0.00%'                 # Changes percent below high column format to percentage.

for row_number in range(1, sheet_output.max_row+1):                             # Loops through all rows in the output file.
    for col_number in range(1, sheet_output.max_column+1):                      # Loops through all columns in the output file.
        cell_obj = sheet_output.cell(row=row_number, column=col_number)         # Identifies output cells using 'row_number and 'col_number' variables.
        cell_obj.alignment = Alignment(horizontal="center")                     # Centers the horizontal alignment of the data.  
    
#------------------------------------------#
#-------- EXPORT NEW TICKERS FILE ---------#
#------------------------------------------#

wb_output.save('C://StockAnalysis.xlsx')                                        # Output file path.


print("Analysis complete. Data available in StockAnalysis.xlsx.")               # Confirmation message.
